import openpyxl

#reading data from excel
path="C:\Users\Kuba\PycharmProjects\python_selenium\excell.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
# sheet = workbook.get_sheet_by_name("Sheet1")
rows = sheet.max_row
col = sheet.max_column

for r in range(1, rows+1):
    for c in range(1, col+1):
        print(sheet.cell(row=r, column=c).value)


#writing data to excel
path="C:\Users\Kuba\PycharmProjects\python_selenium\excell.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
for r in range(1, 6):
    for c in range(1, 4):
        sheet.cell(row=r, column=c).value="welcome"
workbook.save(path)

#data driven
